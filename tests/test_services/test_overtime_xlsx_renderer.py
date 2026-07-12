from datetime import date, datetime, time, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.services.overtime.xlsx_renderer import OvertimeXlsxRenderer, safe_excel_text


def context(rows=None):
    rows = rows if rows is not None else [{
        "technician": "+Técnico Ñ", "project": "@Proyecto", "work_date_value": date(2026, 7, 8),
        "entry_time_value": time(7), "break_start_time_value": None,
        "break_end_time_value": None, "exit_time_value": time(16, 30),
        "activity": "  =SUM(A1:A2) actividad extensa con Unicode áéíóú",
        "worked_minutes": 1530, "regular_minutes": 480, "overtime_minutes": 1050,
        "status": "Cancelada",
    }]
    count = len(rows)
    return {
        "rows": rows,
        "technician_totals": ([{
            "technician": "+Técnico Ñ", "count": count, "worked_minutes": 1530,
            "regular_minutes": 480, "overtime_minutes": 1050,
        }] if rows else []),
        "grand_total": {"count": count, "worked_minutes": 1530 if rows else 0,
                        "regular_minutes": 480 if rows else 0,
                        "overtime_minutes": 1050 if rows else 0},
        "date_from": "2026-07-01", "date_to": "2026-07-08",
        "generated_at": "2026-07-08 10:00", "company_name": "=Compañía",
        "status_filter": "Todos", "technician_filter": "@Seleccionado",
    }


def open_rendered(value):
    payload = OvertimeXlsxRenderer().render(value)
    assert payload.startswith(b"PK")
    return load_workbook(BytesIO(payload), data_only=False)


def test_xlsx_workbook_has_stable_sheets_types_formats_totals_and_security():
    workbook = open_rendered(context())
    assert workbook.sheetnames == ["Solicitudes", "Resumen por técnico", "Resumen general"]
    sheet = workbook["Solicitudes"]
    assert sheet["B2"].value == "'=Compañía"
    assert sheet["A9"].value == "Técnico"
    assert sheet["L9"].value == "Estado"
    assert sheet.freeze_panes == "A10"
    assert sheet.auto_filter.ref == "A9:L10"
    assert isinstance(sheet["C10"].value, (date, datetime))
    assert sheet["C10"].number_format == "yyyy-mm-dd"
    assert sheet["D10"].number_format == "hh:mm"
    assert sheet["E10"].value is None and sheet["F10"].value is None
    assert sheet["I10"].number_format == "[h]:mm"
    assert sheet["I10"].value == timedelta(minutes=1530)
    for cell in (sheet["A10"], sheet["B10"], sheet["H10"]):
        assert cell.data_type == "s"
        assert cell.value.lstrip().startswith("'")
    assert sheet["L10"].value == "Cancelada"
    summary = workbook["Resumen por técnico"]
    assert summary["B2"].value == 1
    assert summary["C2"].number_format == "[h]:mm"
    general = workbook["Resumen general"]
    assert general["B5"].value == 1
    assert general["B6"].number_format == "[h]:mm"
    assert "no constituyen una nómina pagable" in general["B9"].value
    workbook.close()


def test_xlsx_empty_result_keeps_three_valid_sheets_headers_and_zero_totals():
    workbook = open_rendered(context([]))
    assert workbook["Solicitudes"].max_row == 9
    assert workbook["Solicitudes"].auto_filter.ref == "A9:L9"
    assert workbook["Resumen por técnico"].max_row == 1
    assert workbook["Resumen general"]["B5"].value == 0
    workbook.close()


def test_safe_excel_text_only_prefixes_formula_like_values_after_leading_spaces():
    for value in ("=SUM(A1:A2)", "+cmd", "-1+1", "@formula", "  =formula"):
        assert safe_excel_text(value).lstrip().startswith("'")
    assert safe_excel_text("Texto normal") == "Texto normal"
