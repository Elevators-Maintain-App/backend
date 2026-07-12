from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def safe_excel_text(value: object) -> str:
    text = str(value)
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


class OvertimeXlsxRenderer:
    DURATION_FORMAT = "[h]:mm"
    HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

    def render(self, context: dict) -> bytes:
        workbook = Workbook()
        requests_sheet = workbook.active
        requests_sheet.title = "Solicitudes"
        technician_sheet = workbook.create_sheet("Resumen por técnico")
        general_sheet = workbook.create_sheet("Resumen general")

        self._write_requests(requests_sheet, context)
        self._write_technician_summary(technician_sheet, context)
        self._write_general_summary(general_sheet, context)

        buffer = BytesIO()
        try:
            workbook.save(buffer)
            return buffer.getvalue()
        finally:
            workbook.close()
            buffer.close()

    @classmethod
    def _write_requests(cls, sheet, context: dict) -> None:
        metadata = (
            ("Reporte de solicitudes de horas extra", ""),
            ("Compañía", safe_excel_text(context["company_name"])),
            ("Período", f'{context["date_from"]} a {context["date_to"]}'),
            ("Generado en Panamá", context["generated_at"]),
            ("Estado", safe_excel_text(context["status_filter"])),
            ("Técnico", safe_excel_text(context["technician_filter"])),
            ("Cantidad total", context["grand_total"]["count"]),
        )
        for row in metadata:
            sheet.append(row)
        sheet["A1"].font = Font(bold=True, size=14)

        headers = (
            "Técnico", "Proyecto", "Fecha", "Entrada", "Inicio receso", "Fin receso",
            "Salida", "Actividad", "Tiempo trabajado", "Tiempo regular", "Horas extra", "Estado",
        )
        header_row = 9
        sheet.append(())
        sheet.append(headers)
        cls._style_header(sheet, header_row)

        for item in context["rows"]:
            sheet.append((
                safe_excel_text(item["technician"]), safe_excel_text(item["project"]),
                item["work_date_value"], item["entry_time_value"],
                item["break_start_time_value"], item["break_end_time_value"],
                item["exit_time_value"], safe_excel_text(item["activity"]),
                item["worked_minutes"] / 1440, item["regular_minutes"] / 1440,
                item["overtime_minutes"] / 1440, safe_excel_text(item["status"]),
            ))
            row_number = sheet.max_row
            sheet.cell(row_number, 3).number_format = "yyyy-mm-dd"
            for column in (4, 5, 6, 7):
                sheet.cell(row_number, column).number_format = "hh:mm"
            for column in (9, 10, 11):
                sheet.cell(row_number, column).number_format = cls.DURATION_FORMAT
            sheet.cell(row_number, 8).alignment = Alignment(wrap_text=True, vertical="top")

        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:L{max(header_row, sheet.max_row)}"
        widths = (24, 24, 12, 10, 14, 12, 10, 42, 17, 15, 13, 20)
        for column, width in enumerate(widths, 1):
            sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width

    @classmethod
    def _write_technician_summary(cls, sheet, context: dict) -> None:
        headers = ("Técnico", "Cantidad de solicitudes", "Tiempo trabajado", "Tiempo regular", "Horas extra")
        sheet.append(headers)
        cls._style_header(sheet, 1)
        for total in context["technician_totals"]:
            sheet.append((
                safe_excel_text(total["technician"]), total["count"],
                total["worked_minutes"] / 1440, total["regular_minutes"] / 1440,
                total["overtime_minutes"] / 1440,
            ))
            for column in (3, 4, 5):
                sheet.cell(sheet.max_row, column).number_format = cls.DURATION_FORMAT
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:E{max(1, sheet.max_row)}"
        for column, width in enumerate((28, 23, 18, 16, 14), 1):
            sheet.column_dimensions[sheet.cell(1, column).column_letter].width = width

    @classmethod
    def _write_general_summary(cls, sheet, context: dict) -> None:
        total = context["grand_total"]
        rows = (
            ("Resumen general", ""),
            ("Período", f'{context["date_from"]} a {context["date_to"]}'),
            ("Estado", safe_excel_text(context["status_filter"])),
            ("Técnico", safe_excel_text(context["technician_filter"])),
            ("Cantidad total", total["count"]),
            ("Tiempo trabajado total", total["worked_minutes"] / 1440),
            ("Tiempo regular total", total["regular_minutes"] / 1440),
            ("Horas extra totales", total["overtime_minutes"] / 1440),
            ("Nota", "Los totales representan solicitudes filtradas y no constituyen una nómina pagable"),
        )
        for row in rows:
            sheet.append(row)
        sheet["A1"].font = Font(bold=True, size=14)
        for row in (6, 7, 8):
            sheet.cell(row, 2).number_format = cls.DURATION_FORMAT
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 82
        sheet["B9"].alignment = Alignment(wrap_text=True)

    @classmethod
    def _style_header(cls, sheet, row: int) -> None:
        for cell in sheet[row]:
            cell.font = Font(bold=True)
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(vertical="center")
